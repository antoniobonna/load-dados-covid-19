# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 01:51:25 2019

@author: Antonio
"""

import os
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import psycopg2
import credentials

countries = ['Brazil','United States','Italy','Spain','United Kingdom','Germany','France','Russia']
file = 'tabelas_email.xlsx'
indir = '/home/ubuntu/dump/dados_covid_19/'
outdir = '/home/ubuntu/scripts/load-dados-covid-19/'

query_highlights = '''SELECT "#", country, cases, total_rate, population_rate, deaths, deaths_per_million, death_rate, recovered, tests, tests_per_thousand
	FROM covid_19_dw.vw_highlights ORDER BY 1 LIMIT 20'''

query_country = '''SELECT date, cases, new_cases, cases_evolution, deaths, new_deaths, deaths_evolution, death_rate
	FROM covid_19_dw.vw_world_data
	WHERE country = '{}' 
	ORDER BY date OFFSET 1'''

query_brazil_states = '''SELECT "#", state, cases, total_rate, population_rate, deaths, deaths_per_million, death_rate, recovered,
    tests, tests_per_thousand
    FROM covid_19_dw.vw_brazil_states ORDER BY 1 LIMIT 20'''

query_brazil_cities = '''SELECT "#", city, cases, total_rate, population_rate, deaths, deaths_per_million, death_rate
	FROM covid_19_dw.vw_brazil_cities ORDER BY 1 LIMIT 20'''

query_local_occupation = '''SELECT CASE WHEN local = 'Rio Grande do Sul' THEN 'Rio Grande do Sul**'
	WHEN local = 'São Paulo - SP' THEN 'São Paulo - SP*'
    WHEN local in ('Ceará','Rondônia') THEN CONCAT(local,'***')
    WHEN local = 'Rio de Janeiro - RJ' THEN 'Rio de Janeiro - RJ****'
    WHEN local = 'São Paulo' THEN 'São Paulo*****'
    WHEN local = 'Amazonas' THEN 'Amazonas******'
    ELSE local END as local, 
	bed_number, inpatients, occupation, inpatients_growth, status
	FROM covid_19_dw.vw_local_occupation ORDER BY occupation DESC'''

query_lockdown = '''SELECT t.country
     , t.start_date
     , t.end_date
FROM covid_19.lockdown t WHERE country in ('Austria','France','Germany','Italy','Spain','Iran')
ORDER BY t.country'''

DATABASE, HOST, USER, PASSWORD = credentials.setDatabaseLogin()

### conecta no banco de dados
db_conn = psycopg2.connect("dbname='{}' user='{}' host='{}' password='{}'".format(DATABASE, USER, HOST, PASSWORD))
cursor = db_conn.cursor()
print('Connected to the database')

wb = load_workbook(filename=indir+file, read_only=False)

sheet = wb['Highlights']

cursor.execute(query_highlights)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = i+2, column = j+1).value = result[i][j]

for country in countries:
    sheet = wb[country]
    cursor.execute(query_country.format(country))
    result = [item for item in cursor.fetchall()]
    for i in range(len(result)):
        for j in range(len(result[i])):
            sheet.cell(row = i+3, column = j+1).value = result[i][j]

sheet = wb['Brazil States']

cursor.execute(query_brazil_states)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = i+2, column = j+1).value = result[i][j]

sheet = wb['Brazil Cities']

cursor.execute(query_brazil_cities)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = i+2, column = j+1).value = result[i][j]

sheet = wb['Beds Occupation']

cursor.execute(query_local_occupation)
result = [item for item in cursor.fetchall()]

for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = i+2, column = j+1).value = result[i][j]
        if result[i][j] == 'Sob Controle':
            # sheet.cell(row = i+2, column = j+1).fill = PatternFill(start_color="70AD47", fill_type = "solid")
            sheet.cell(row = i+2, column = j+1).font = Font(color="548235", bold=True)
        elif result[i][j] == 'Alerta':
            # sheet.cell(row = i+2, column = j+1).fill = PatternFill(start_color="FFD966", fill_type = "solid")
            sheet.cell(row = i+2, column = j+1).font = Font(color="FFC000", bold=True)
        elif result[i][j] == 'Caos':
            # sheet.cell(row = i+2, column = j+1).fill = PatternFill(start_color="FF5050", fill_type = "solid")
            sheet.cell(row = i+2, column = j+1).font = Font(color="FF5050", bold=True)

sheet = wb['Lockdown']
cursor.execute(query_lockdown)
result = [item for item in cursor.fetchall()]
for i in range(len(result)):
    for j in range(len(result[i])):
        sheet.cell(row = i+3, column = j+1).value = result[i][j]

wb.save(outdir+file)
wb.close()
cursor.close()
db_conn.close()